# !/usr/bin/python
'''
██╗  ██╗ ██████╗ ████████╗██╗    ██╗ ██████╗ ██████╗ 
╚██╗██╔╝██╔═══██╗╚══██╔══╝██║    ██║██╔═══██╗██╔══██╗
 ╚███╔╝ ██║   ██║   ██║   ██║ █╗ ██║██║   ██║██║  ██║
 ██╔██╗ ██║   ██║   ██║   ██║███╗██║██║   ██║██║  ██║
██╔╝ ██╗╚██████╔╝   ██║   ╚███╔███╔╝╚██████╔╝██████╔╝
╚═╝  ╚═╝ ╚═════╝    ╚═╝    ╚══╝╚══╝  ╚═════╝ ╚═════╝ 
                                                     
'''

#----Imports----#
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import argparse
import sys
import os
import time
timestr = time.strftime("%Y%m%d-%H%M%S")

#----Arg parsing----#
parser = argparse.ArgumentParser(description='Example: Excel.py 19011200020001')
parser.add_argument("-s","--serial",metavar='',help="Serialnumber from Matas.",required=True)
group = parser.add_mutually_exclusive_group()
group.add_argument('-q','--quiet',action='store_true',help='App is running quiet')
group.add_argument('-v','--verbose',action='store_true',help='App is running verbose')
args = parser.parse_args()

#----Arg mapping----#
matasqr = args.serial

#----Workbook and sheet loading----#
wb = load_workbook(filename = 'Workbook.xlsx')
sheet = wb['Sheet1']
sheet._number_formats = '0.00E+00'
row_count = sheet.max_row + 1
column_count = sheet.max_column + 1

#----Functions----#

#----Itterate through all Rows and columns----#
def itterate():
    for i in range(1,row_count+1):
        print(sheet.cell(row=i,column=1).value)
    for i in range(1,column_count+1):
        print(sheet.cell(row=1,column=i).value)

#----Find an empty cell and link the matasqr to it, return the value of the cell left of it. Continue to execute programming with data from left row.----#
def linkserial():
    print("Using Serial: " + args.serial)
    for cell in sheet["b"]:
        if cell.value is None:
            if args.verbose:
                print("Found empty cell %d, writing to it." % cell.row)
                print("Linked serial: %d " % sheet.cell(row=cell.row,column=1).value)
            clientlink = sheet.cell(row=cell.row,column=1).value              
            sheet.cell(row=cell.row,column=2).value = matasqr
            print("Linked to: %d" % clientlink)
            os.system('cmd /c "C:\Programming\python\Python-Projects\Excel_read_serialnumber\FW\MCU\FirmwareflashMCU_Serial.bat %d"' % clientlink)
            break
    else:
        print("Couldn't link serials... Exiting...")
        sys.exit()
        

#----Main----#
if __name__ == "__main__":
    
    if args.quiet:
        print("--Running Quiet--")
        linkserial()
    elif args.verbose:
        print("--Running Verbose--")
        print("Sheet names:")
        print(wb.sheetnames)
        print("Row count: %d" % row_count)
        print("Column count: %d " % column_count)
        print("Sheet formatting: %s " % sheet._number_formats)
        print('Saving to: Workbook.xlsx')
        print('Creating Backup: Workbook %s.xlsx' % timestr)
        linkserial()
    elif args.serial:
        linkserial()
    else:
        print("You broke it.")
        sys.exit()


    #----Save changes to different .xlsx to prevent overriding data----#
    wb.save('Workbook.xlsx')
    wb.save('Workbook %s.xlsx' % timestr)